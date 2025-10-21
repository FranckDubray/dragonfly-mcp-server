from __future__ import annotations
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from .mapping import detect_headers, build_header_maps, normalize
from .formatters import build_style_bundle, apply_row_style, apply_per_column_styles
from .utils import parse_date_input, compute_row_range


def _resolve_sheet(wb, sheet_name) -> Worksheet:
    if isinstance(sheet_name, int):
        return wb.worksheets[sheet_name]
    if sheet_name is None:
        return wb.active
    return wb[sheet_name]


def _coerce_row_values(row_data: Dict[str, Any] | None, norm_to_real: Dict[str, str], parse_cfg: dict | None) -> Dict[str, Any]:
    if not row_data:
        return {}
    out: Dict[str, Any] = {}
    date_mode = (parse_cfg or {}).get("date_input_format", "auto")
    for k, v in row_data.items():
        # map normalized name to real header if possible
        real = k
        nk = normalize(k)
        if nk in norm_to_real:
            real = norm_to_real[nk]
        # date parsing only on strings; leave others as is
        vv = parse_date_input(v, date_mode)
        out[real] = vv
    return out


def insert_row(ws: Worksheet, position: int, headers: list[str], values: Dict[str, Any]) -> int:
    # Insert after header row (row 1 is headers). position 1 means row index 2.
    target_idx = position + 1
    ws.insert_rows(target_idx)
    # write cells according to headers order
    for i, h in enumerate(headers, start=1):
        ws.cell(row=target_idx, column=i, value=values.get(h))
    return target_idx


def update_row(ws: Worksheet, position: int, headers: list[str], values: Dict[str, Any]) -> int:
    target_idx = position + 1
    max_row = ws.max_row
    if target_idx > max_row:
        raise ValueError("update_row: position beyond last row")
    for i, h in enumerate(headers, start=1):
        if h in values:
            ws.cell(row=target_idx, column=i, value=values[h])
    return target_idx


def delete_row(ws: Worksheet, position: int) -> int:
    target_idx = position + 1
    ws.delete_rows(target_idx)
    return target_idx


def apply_formatting(ws: Worksheet, row_idx: int, headers: list[str], row_fmt_cfg: dict | None, per_col_cfg: dict | None, header_map_norm_to_real: dict):
    # Build bundles once
    row_bundle = build_style_bundle(row_fmt_cfg)
    cells = [ws.cell(row=row_idx, column=i) for i in range(1, len(headers) + 1)]
    # Row-level
    apply_row_style(cells, row_bundle)
    # Per-column
    cells_by_name = {h: ws.cell(row=row_idx, column=i+1-1) for i, h in enumerate(headers, start=1)}
    apply_per_column_styles(cells_by_name, per_col_cfg, header_map_norm_to_real)


def operate(
    operation: str,
    file_path: str,
    sheet_name,
    position: int | None,
    row_data: Dict[str, Any] | None,
    row_format: dict | None,
    columns_format: dict | None,
    parse: dict | None,
):
    wb = load_workbook(filename=file_path)
    ws = _resolve_sheet(wb, sheet_name)

    # headers from first row
    headers = detect_headers([c.value for c in ws[1]])
    norm_to_real, real_to_idx = build_header_maps(headers)

    result = {
        "operation": operation,
        "sheet_name": ws.title,
        "table_columns": headers,
        "warnings": [],
    }

    if operation == "list_backups":
        # Placeholder: backups handled in api
        result["success"] = True
        result["backups"] = []
        return result

    if operation == "restore_backup":
        # Placeholder: handled in api
        result["success"] = True
        return result

    if operation in {"insert_row", "update_row", "delete_row"} and (position is None or position < 1):
        raise ValueError("position must be provided and >= 1")

    row_idx = None

    if operation == "insert_row":
        values = _coerce_row_values(row_data, norm_to_real, parse)
        row_idx = insert_row(ws, position, headers, values)
        apply_formatting(ws, row_idx, headers, row_format, columns_format, norm_to_real)
    elif operation == "update_row":
        values = _coerce_row_values(row_data, norm_to_real, parse)
        row_idx = update_row(ws, position, headers, values)
        apply_formatting(ws, row_idx, headers, row_format, columns_format, norm_to_real)
    elif operation == "delete_row":
        row_idx = delete_row(ws, position)
    else:
        raise ValueError("Unsupported operation in core")

    # Save
    wb.save(file_path)

    from .utils import compute_row_range
    if row_idx:
        result["formatted_range"] = compute_row_range(len(headers), row_idx)
    result["success"] = True
    result["row_index_excel"] = row_idx
    result["total_rows_after"] = ws.max_row - 1 if ws.max_row >= 1 else 0
    return result
