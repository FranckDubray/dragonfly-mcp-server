from __future__ import annotations
from typing import Any
from datetime import datetime
from openpyxl.utils import get_column_letter
import os
from pathlib import Path

# Robust project root detection (final):
# Priority chain:
# 0) ENV DGY_PROJECT_ROOT or PROJECT_ROOT (if valid dir)
# 1) CWD if contains 'docs/office' or 'src'
# 2) Nearest ancestor that contains both 'src' and 'docs/office'
# 3) Parent of nearest 'src' ancestor
# 4) Fallback: 3 levels up from this file
_env = os.getenv('DGY_PROJECT_ROOT') or os.getenv('PROJECT_ROOT')
_root_path: Path | None = None
if _env and os.path.isdir(_env):
    _root_path = Path(_env).resolve()
else:
    # 1) Try CWD
    cwd = Path(os.getcwd()).resolve()
    if (cwd / 'docs' / 'office').is_dir() or (cwd / 'src').is_dir():
        _root_path = cwd
    if _root_path is None:
        here = Path(__file__).resolve()
        # 2) Both src and docs/office present
        for p in here.parents:
            if (p / 'src').is_dir() and (p / 'docs' / 'office').is_dir():
                _root_path = p
                break
        # 3) Parent of nearest 'src'
        if _root_path is None:
            for p in here.parents:
                if p.name == 'src':
                    _root_path = p.parent
                    break
        # 4) Fallback
        if _root_path is None:
            _root_path = here.parents[3]

ROOT = str(_root_path)


def safe_join(path: str) -> str:
    # Allow absolute paths only if inside ROOT; otherwise join ROOT + path
    if os.path.isabs(path):
        p = os.path.abspath(path)
    else:
        p = os.path.abspath(os.path.join(ROOT, path))
    if not p.startswith(ROOT):
        raise ValueError("file_path outside project root")
    return p


DATE_PATTERNS = {
    "YYYY-MM-DD": "%Y-%m-%d",
    "DD/MM/YYYY": "%d/%m/%Y",
    "MM/DD/YYYY": "%m/%d/%Y",
}


def parse_date_input(value: Any, mode: str = "auto"):
    if not isinstance(value, str):
        return value
    if mode == "auto":
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except Exception:
                pass
        return value
    fmt = DATE_PATTERNS.get(mode)
    if not fmt:
        return value
    try:
        return datetime.strptime(value.strip(), fmt)
    except Exception:
        return value


def get_table_headers(ws) -> list[str]:
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val is None or str(val).strip() == "":
            break
        headers.append(str(val))
    return headers


def column_letter(idx: int) -> str:
    return get_column_letter(idx)


def compute_row_range(col_count: int, row_idx: int) -> str:
    last = column_letter(col_count)
    return f"A{row_idx}:{last}{row_idx}"
