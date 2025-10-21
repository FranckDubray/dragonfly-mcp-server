from __future__ import annotations
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, PatternFill, Border, Side

_CACHE: dict[str, object] = {}


def _key(prefix: str, d: dict | None) -> str:
    if not d:
        return prefix
    items = sorted((k, v) for k, v in d.items() if v is not None)
    return prefix + "|" + ";".join(f"{k}={v}" for k, v in items)


def build_font(cfg: dict | None) -> XLFont | None:
    if not cfg:
        return None
    key = _key("font", cfg)
    if key in _CACHE:
        return _CACHE[key]
    
    # Build kwargs only for non-None values
    kwargs = {}
    if cfg.get("name"):
        kwargs["name"] = cfg["name"]
    if cfg.get("size") is not None:
        kwargs["size"] = float(cfg["size"])
    if cfg.get("bold") is not None:
        kwargs["bold"] = bool(cfg["bold"])
    if cfg.get("italic") is not None:
        kwargs["italic"] = bool(cfg["italic"])
    if cfg.get("underline"):
        kwargs["underline"] = "single"
    if cfg.get("color"):
        kwargs["color"] = cfg["color"].replace("#", "")
    
    f = XLFont(**kwargs)
    _CACHE[key] = f
    return f


def build_alignment(cfg: dict | None) -> XLAlignment | None:
    if not cfg:
        return None
    key = _key("aln", cfg)
    if key in _CACHE:
        return _CACHE[key]
    
    # Build kwargs only for non-None values to avoid openpyxl type issues
    kwargs = {}
    if cfg.get("horizontal"):
        kwargs["horizontal"] = cfg["horizontal"]
    if cfg.get("vertical"):
        kwargs["vertical"] = cfg["vertical"]
    if cfg.get("wrap_text") is not None:
        kwargs["wrapText"] = bool(cfg["wrap_text"])
    if cfg.get("indent") is not None:
        kwargs["indent"] = int(cfg["indent"])
    
    a = XLAlignment(**kwargs)
    _CACHE[key] = a
    return a


def build_fill(cfg: dict | None) -> PatternFill | None:
    if not cfg or not cfg.get("background_color"):
        return None
    key = _key("fill", cfg)
    if key in _CACHE:
        return _CACHE[key]
    col = cfg.get("background_color").replace("#", "")
    pf = PatternFill(fill_type="solid", fgColor=col)
    _CACHE[key] = pf
    return pf


_BORDER_STYLES = {"thin", "medium", "thick", "none"}

def build_border(cfg: dict | None) -> Border | None:
    if not cfg:
        return None
    style = (cfg.get("style") or "thin").lower()
    if style not in _BORDER_STYLES:
        style = "thin"
    if style == "none":
        return None
    key = _key("bdr", cfg)
    if key in _CACHE:
        return _CACHE[key]
    color = (cfg.get("color") or "#000000").replace("#", "")
    side = Side(style=style, color=color)
    b = Border(left=side, right=side, top=side, bottom=side)
    _CACHE[key] = b
    return b
